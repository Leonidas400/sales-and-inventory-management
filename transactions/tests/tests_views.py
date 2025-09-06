import json
from io import BytesIO
from openpyxl import load_workbook
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import User
from ..models import Sale, Purchase, SaleDetail
from store.models import Item, Category
from accounts.models import Customer, Vendor, Profile

class SaleViewsTest(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username='teste', password='senha123')
        cls.superuser = User.objects.create_superuser(username='superteste', password='senha123')
        cls.customer = Customer.objects.create(first_name='Cliente Teste')
        category = Category.objects.create(name='Cat Teste')
        cls.item = Item.objects.create(name='Item Teste', price=100, quantity=20, category=category)
        cls.sale = Sale.objects.create(customer=cls.customer, grand_total=100)

    def test_sale_list_view_requires_login(self):
        response = self.client.get(reverse('saleslist'))
        self.assertRedirects(response, f"{reverse('user-login')}?next={reverse('saleslist')}")

    def test_sale_list_view_loads_correctly(self):
        self.client.login(username='teste', password='senha123')
        response = self.client.get(reverse('saleslist'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'transactions/sales_list.html')

    def test_sale_create_view_post_success(self):
        self.client.login(username='teste', password='senha123')
        initial_item_quantity = self.item.quantity
        
        sale_data = {
            'customer': self.customer.pk,
            'sub_total': 200, 'grand_total': 200,
            'amount_paid': 200, 'amount_change': 0,
            'items': [{'id': self.item.pk, 'price': 100, 'quantity': 2, 'total_item': 200}]
        }
        
        response = self.client.post(
            reverse('sale-create'),
            data=json.dumps(sale_data),
            content_type='application/json',
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'success')
        self.item.refresh_from_db()
        self.assertEqual(self.item.quantity, initial_item_quantity - 2)

    def test_sale_create_view_post_insufficient_stock(self):
        self.client.login(username='teste', password='senha123')
        
        sale_data = {
            'customer': self.customer.pk, 'sub_total': 20000, 'grand_total': 20000,
            'amount_paid': 20000, 'amount_change': 0,
            'items': [{'id': self.item.pk, 'price': 100, 'quantity': 99, 'total_item': 9900}]
        }
        
        response = self.client.post(
            reverse('sale-create'), data=json.dumps(sale_data), content_type='application/json',
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        
        self.assertEqual(response.status_code, 400)
        self.assertIn('Not enough stock', response.json()['message'])

    def test_sale_delete_view_permission_denied_for_normal_user(self):
        self.client.login(username='teste', password='senha123')
        response = self.client.get(reverse('sale-delete', args=[self.sale.pk]))
        self.assertEqual(response.status_code, 403)


class PurchaseViewsTest(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username='teste', password='senha123')
        cls.superuser = User.objects.create_superuser(username='superteste', password='senha123')
        cls.vendor = Vendor.objects.create(name="Fornecedor Teste")
        category = Category.objects.create(name='Cat Teste')
        cls.item = Item.objects.create(name='Item Teste', price=100, quantity=20, category=category)
        cls.purchase = Purchase.objects.create(item=cls.item, vendor=cls.vendor, quantity=5, price=100)

    def test_purchase_list_view_loads_correctly(self):
        self.client.login(username='teste', password='senha123')
        response = self.client.get(reverse('purchaseslist'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'transactions/purchases_list.html')

    def test_purchase_delete_view_by_superuser(self):
        self.client.login(username='superteste', password='senha123')
        initial_count = Purchase.objects.count()
        response = self.client.post(reverse('purchase-delete', args=[self.purchase.pk]))
        self.assertEqual(Purchase.objects.count(), initial_count - 1)
        self.assertRedirects(response, reverse('purchaseslist'))


class ExportViewsTest(TestCase):

    @classmethod
    def setUpTestData(cls):
        customer = Customer.objects.create(
            first_name='Cliente Export', 
            phone='11969420872'
        )
        Sale.objects.create(customer=customer, grand_total=500)

    def test_export_sales_to_excel_view(self):
        response = self.client.get(reverse('sales-export'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, 'ID')
        self.assertEqual(worksheet['C2'].value, '11969420872')
        self.assertEqual(worksheet.max_row, 2)